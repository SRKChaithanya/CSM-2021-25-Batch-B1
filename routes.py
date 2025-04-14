import pandas as pd
from flask import render_template, request, flash, redirect, url_for, send_file, jsonify
from flask_login import login_user, logout_user, login_required, current_user
from app import app, db
from models import User, Department, Subject, Template, Question, Regulation, Image, SelectedQuestion
from forms import LoginForm, DepartmentForm
import json
import io
import datetime
import math
import pandas as pd
from utils import generate_pdf
from image_utils import extract_image_from_excel, get_image_base64
from openpyxl import load_workbook
from jinja2 import Environment
from sqlalchemy import or_


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user is None or not user.check_password(form.password.data):
            flash('Invalid username or password', 'danger')
            return redirect(url_for('login'))

        login_user(user)
        next_page = request.args.get('next')
        # Simple security check for next parameter
        if not next_page or next_page.startswith('//'):
            next_page = url_for('index')
        return redirect(next_page)

    return render_template('login.html', form=form)


app.jinja_env.filters['char'] = lambda n: chr(n)
app.jinja_env.filters['sum'] = lambda l: sum(l) if l else 0


@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    # Count departments
    departments_count = Department.query.count()
    
    # Count regulations
    regulations_count = Regulation.query.count()
    
    # Get all departments for filter
    all_departments = Department.query.all()
    
    # Get all regulations for filter
    all_regulations = Regulation.query.all()
    
    # Exam types list
    exam_types = ['CIE-I', 'CIE-II', 'SEE']
    
    # Paper types list
    paper_types = ['Regular', 'Supplementary', 'Regular/Supplementary']
    
    # For chart data - analyze question paper composition
    # Get all selected questions to analyze composition by CO and BL
    selected_questions = SelectedQuestion.query.all()
    
    # Initialize data structures for chart
    co_distribution = {}
    bl_distribution = {}
    
    for sq in selected_questions:
        # Get the actual question
        question = Question.query.get(sq.question_id)
        if question:
            # CO distribution
            if question.co in co_distribution:
                co_distribution[question.co] += 1
            else:
                co_distribution[question.co] = 1
                
            # BL distribution
            if question.bl in bl_distribution:
                bl_distribution[question.bl] += 1
            else:
                bl_distribution[question.bl] = 1
    
    # Convert to format needed for Chart.js
    co_labels = list(co_distribution.keys())
    co_values = list(co_distribution.values())
    
    bl_labels = list(bl_distribution.keys())
    bl_values = list(bl_distribution.values())
    
    return render_template('index.html', 
                          departments_count=departments_count,
                          regulations_count=regulations_count,
                          all_departments=all_departments,
                          all_regulations=all_regulations,
                          exam_types=exam_types,
                          paper_types=paper_types,
                          co_labels=co_labels,
                          co_values=co_values,
                          bl_labels=bl_labels,
                          bl_values=bl_values)


@app.route('/departments')
@login_required
def departments():
    departments = Department.query.all()
    return render_template('departments.html', departments=departments)


@app.route('/department/add', methods=['GET', 'POST'])
@login_required
def add_department():
    form = DepartmentForm()
    if form.validate_on_submit():
        department = Department(name=form.name.data, dept_id=form.dept_id.data)
        db.session.add(department)
        db.session.commit()
        flash('Department added successfully!', 'success')
        return redirect(url_for('departments'))
    return render_template('department_form.html', form=form)


@app.route('/create_template', methods=['GET', 'POST'])
@login_required
def create_template():
    regulation_id = request.args.get('regulation_id')
    exam_type = request.args.get('exam_type')

    if not regulation_id or not exam_type:
        flash('Regulation and exam type are required', 'error')
        return redirect(url_for('regulations'))

    regulation = Regulation.query.get_or_404(regulation_id)

    if request.method == 'POST':
        try:
            # Get header information from form
            header = {
                'institution': request.form['institution'],
                'autonomous': request.form.get('autonomous', 'AUTONOMOUS'),
                'time_hours': int(request.form['time_hours']),
                'max_marks': int(request.form['max_marks'])
            }

            # Get Part A data
            partA_data = json.loads(request.form['partA'])
            if not partA_data:
                flash('At least one Part A question is required', 'error')
                return redirect(request.url)

            # Get Part B data
            partB_data = json.loads(request.form['partB'])
            if not partB_data:
                flash('At least one Part B question is required', 'error')
                return redirect(request.url)

            # Create new template
            template = Template(name=f"{regulation.name} - {exam_type}",
                                exam_type=exam_type,
                                header=header,
                                partA=partA_data,
                                partB=partB_data,
                                regulation_id=regulation_id)

            db.session.add(template)
            db.session.commit()

            flash('Template created successfully!', 'success')
            return redirect(url_for('regulations'))

        except Exception as e:
            flash(f'Error creating template: {str(e)}', 'error')
            return redirect(request.url)

    return render_template('create_template.html',
                           regulation=regulation,
                           exam_type=exam_type)


@app.route('/upload_questions', methods=['GET', 'POST'])
def upload_questions():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)

        if not file.filename.endswith('.xlsx'):
            flash('Please upload an Excel file', 'error')
            return redirect(request.url)

        try:
            # Read Excel file
            df = pd.read_excel(file, sheet_name=None)

            for unit, sheet in df.items():
                if not all(col in sheet.columns
                           for col in ['Question', 'Marks', 'CO', 'BL']):
                    flash(f'Missing required columns in unit {unit}', 'error')
                    continue

                for _, row in sheet.iterrows():
                    question = Question(unit=int(unit.split()[-1]),
                                        question_text=row['Question'],
                                        marks=row['Marks'],
                                        co=row['CO'],
                                        bl=row['BL'])
                    db.session.add(question)

            db.session.commit()
            flash('Questions uploaded successfully!', 'success')

        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'error')

    return render_template('question_bank.html',
                           questions=Question.query.all())


@app.route('/regulations')
@login_required
def regulations():
    regulations = Regulation.query.all()
    exam_types = ['CIE-I', 'CIE-II', 'SEE']
    templates = {
        reg.id: {
            exam_type:
            Template.query.filter_by(regulation_id=reg.id,
                                     exam_type=exam_type).first()
            for exam_type in exam_types
        }
        for reg in regulations
    }
    return render_template('regulations.html',
                           regulations=regulations,
                           templates=templates,
                           exam_types=exam_types)


@app.route('/add_regulation', methods=['POST'])
@login_required
def add_regulation():
    name = request.form.get('name')
    if name:
        regulation = Regulation(name=name)
        db.session.add(regulation)
        db.session.commit()
        flash('Regulation added successfully!', 'success')
    return redirect(url_for('regulations'))


@app.route('/regulation/<int:regulation_id>/rename', methods=['POST'])
@login_required
def rename_regulation(regulation_id):
    regulation = Regulation.query.get_or_404(regulation_id)
    name = request.form.get('name')
    if name:
        regulation.name = name
        db.session.commit()
        flash('Regulation renamed successfully!', 'success')
    return redirect(url_for('regulations'))


@app.route('/regulation/<int:regulation_id>/delete', methods=['POST'])
@login_required
def delete_regulation(regulation_id):
    regulation = Regulation.query.get_or_404(regulation_id)

    # Check if regulation has subjects
    if Subject.query.filter_by(regulation_id=regulation_id).first():
        flash(
            'Remove the subjects assigned to this regulation before deleting this regulation',
            'error')
        return redirect(url_for('regulations'))

    try:
        Template.query.filter_by(regulation_id=regulation_id).delete()
        db.session.delete(regulation)
        db.session.commit()
        flash('Regulation deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting regulation: {str(e)}', 'error')

    return redirect(url_for('regulations'))


@app.route('/create_templates/<int:regulation_id>/<exam_type>')
@login_required
def create_templates(regulation_id, exam_type):
    regulation = Regulation.query.get_or_404(regulation_id)
    template = Template.query.filter_by(regulation_id=regulation_id,
                                        exam_type=exam_type).first()
    return render_template('create_templates.html',
                           regulation=regulation,
                           exam_type=exam_type,
                           template=template)


@app.route('/template/<int:template_id>', methods=['GET'])
@login_required
def view_template(template_id):
    template = Template.query.get_or_404(template_id)
    # Pass the exam_type from the template to ensure all fields are displayed correctly
    return render_template('view_template.html', template=template, exam_type=template.exam_type)


@app.route('/delete_template/<int:template_id>', methods=['POST'])
@login_required
def delete_template(template_id):
    template = Template.query.get_or_404(template_id)
    db.session.delete(template)
    db.session.commit()
    flash('Template deleted successfully', 'success')
    return redirect(url_for('regulations'))


@app.route('/get_questions/<int:unit>/<string:course_code>')
@login_required
def get_questions(unit, course_code):
    marks = request.args.get('marks', type=int)
    dept_id = request.args.get('dept_id')
    exam_type = request.args.get('exam_type', 'SEE')  # Default to SEE if not specified
    apply_fuzzy = request.args.get('apply_fuzzy', 'false') == 'true'
    
    if not dept_id:
        # Try to get department ID from the URL referer
        referer = request.headers.get('Referer', '')
        if 'dept_id=' in referer:
            dept_id = referer.split('dept_id=')[1].split('&')[0]
    
    # Base query with the correct composite foreign key
    query = Question.query.filter_by(unit=unit, subject_course_code=course_code)
    
    if dept_id:
        query = query.filter_by(subject_department_id=dept_id)
        
    if marks:
        query = query.filter_by(marks=marks)

    questions = query.all()

    # Apply fuzzy logic for ranking if requested
    if apply_fuzzy and questions:
        # Get usage history for questions in this unit
        from models import SelectedQuestion
        from sqlalchemy import func

        # Get question usage counts
        usage_counts = {}
        
        # Updated to use composite key fields
        selections_query = SelectedQuestion.query.filter_by(
            subject_course_code=course_code, exam_type=exam_type)
            
        if dept_id:
            selections_query = selections_query.filter_by(subject_department_id=dept_id)
            
        selections = selections_query.all()

        for selection in selections:
            q_id = selection.question_id
            if q_id in usage_counts:
                usage_counts[q_id] += selection.usage_count
            else:
                usage_counts[q_id] = selection.usage_count

        # Calculate fuzzy scores (higher score = less frequently used)
        for q in questions:
            # Formula for fuzzy score:
            # - If never used before: highest score (1.0)
            # - If used before: score decreases with usage count using modified formula:
            # fuzzy_score = 1.0 / sqrt(1.0 + log10(usage_count))
            count = usage_counts.get(q.id, 0)
            if count > 0:
                q.fuzzy_score = 1.0 / math.sqrt(1.0 + math.log10(count))
            else:
                q.fuzzy_score = 1.0  # If never used before, give highest score

        # Sort by fuzzy score (higher score first - less frequently used questions)
        questions.sort(key=lambda x: (-x.fuzzy_score, x.id))

    return jsonify([
        {
            'id': q.id,
            'question_text': q.question_text,
            'marks': q.marks,
            'co': q.co,
            'bl': q.bl,
            'image_data': get_image_base64(q.image.data) if q.image else None,
            'fuzzy_score':
            getattr(q, 'fuzzy_score',
                    1.0)  # Include the fuzzy score in the response
        } for q in questions
    ])


@app.route('/generate_paper/<int:template_id>', methods=['GET', 'POST'])
@login_required
def generate_paper(template_id):
    course_code = request.args.get('course_code')
    dept_id = request.args.get('dept_id')
    exam_type = request.args.get('exam_type')
    paper_type = request.form.get('paper_type', 'Regular')  # Default to Regular

    if not all([course_code, dept_id, exam_type]):
        flash('Missing required parameters', 'error')
        return redirect(url_for('departments'))

    # Get all required objects first
    subject = Subject.query.filter_by(course_code=course_code, department_id=dept_id).first_or_404()
    department = Department.query.get_or_404(dept_id)
    template = Template.query.filter_by(regulation_id=subject.regulation_id,
                                        exam_type=exam_type).first()

    # Get all departments for the dropdown
    all_departments = Department.query.all()

    # Filter departments with same subject, year, semester and regulation
    related_departments = [
        dept for dept in all_departments
        if any(s.year == subject.year and
               s.semester == subject.semester and
               s.regulation_id == subject.regulation_id
               for s in dept.subjects)
    ]

    # Validate all required objects exist
    if not subject:
        flash(f'Subject not found: {course_code}', 'error')
        return redirect(url_for('departments'))
    if not department:
        flash(f'Department not found: {dept_id}', 'error')
        return redirect(url_for('departments'))
    if not template:
        flash(f'Template not found: {template_id}', 'error')
        return redirect(url_for('departments'))

    if not course_code:
        flash('Subject code is required', 'error')
        return redirect(url_for('departments'))

    if not dept_id:
        flash('Department ID is required', 'error')
        return redirect(url_for('departments'))



    subject = Subject.query.filter_by(course_code=course_code, department_id=dept_id).first_or_404()
    department = Department.query.get_or_404(dept_id)
    template = Template.query.filter_by(regulation_id=subject.regulation_id,
                                        exam_type=exam_type).first()

    if not template:
        flash('No template found for this exam type', 'error')
        return redirect(url_for('department_view', dept_id=dept_id))

    if request.method == 'POST':
        try:
            # Add paper_type, month, year and exam_type to template header for PDF generation
            template_header = template.header.copy()
            template_header['paper_type'] = request.form.get(
                'paper_type', 'Regular')
            template_header['exam_month'] = request.form.get(
                'exam_month', 'March')
            template_header['exam_year'] = request.form.get(
                'exam_year', '2025')
            template_header['exam_type'] = exam_type

            # Get selected departments from the form
            selected_departments = request.form.getlist('selected_departments')
            
            # If no departments are selected, use the current department
            if not selected_departments:
                selected_departments = [dept_id]

            # Get selected questions
            part_a_questions = []
            part_b_questions = []

            # Process Part A questions (2 marks questions)
            # Each template position should have its own question
            for q in template.partA:
                q_id = request.form.get(f'partA_{q["qno"]}_{q["sub_qno"]}')

                if q_id:
                    question = Question.query.get(q_id)
                    if question and question.marks == 2:  # Ensure only 2 marks questions
                        # Create a temporary object with the question's attributes
                        temp_question = {
                            'id': question.id,
                            'unit': question.unit,
                            'question_text': question.question_text,
                            'marks': question.marks,
                            'co': question.co,
                            'bl': question.bl,
                            'position': q["qno"],
                            'qno': q["qno"],  # Use question number from form/template 
                            'pdf_qno': request.form.get(f'partA_qno_{q["qno"]}_{q["sub_qno"]}', q["qno"]),  # Get original numbering from form
                            'sub_label': q["sub_qno"] if "sub_qno" in q else q.get("sub_label", chr(97 + int(q["qno"]) - 1) + ')'),
                            'image': question.image
                        }
                        
                        # Add image data if available
                        if temp_question['image']:
                            temp_question['image_data'] = get_image_base64(
                                temp_question['image'].data)
                            
                            # Save resized dimensions from the form if they exist
                            resized_width = request.form.get(f'image_width_{question.id}')
                            resized_height = request.form.get(f'image_height_{question.id}')
                            if resized_width and resized_height:
                                temp_question['resized_width'] = resized_width
                                temp_question['resized_height'] = resized_height
                        
                        # Add the question to the list
                        part_a_questions.append(type('Question', (), temp_question))

            # Process Part B questions (>2 marks questions)
            for q in template.partB:
                q_id = request.form.get(f'partB_{q["qno"]}')
                if q_id:
                    question = Question.query.get(q_id)
                    if question and question.marks > 2:  # Only long answer questions
                        # Create a copy of question attributes to avoid modifying the original
                        question_copy = Question.query.get(question.id)
                        
                        # Get the marks from the form
                        marks = request.form.get(f'marks_{q["qno"]}')
                        marks_value = int(marks) if marks else q.get('marks', 8)
                        
                        # Create a temporary object with the question's attributes
                        temp_question = {
                            'id': question_copy.id,
                            'unit': question_copy.unit,
                            'question_text': question_copy.question_text,
                            'marks': marks_value,
                            'display_marks': marks_value,  # Will be updated for split questions
                            'co': question_copy.co,
                            'bl': question_copy.bl,
                            'qno': q["qno"],  # Include the actual question number from template
                            'pdf_qno': request.form.get(f'partB_qno_{q["qno"]}', q["qno"]), # Get original web interface numbering
                            'image': question_copy.image
                        }
                        
                        # Add image data if available
                        if temp_question['image']:
                            temp_question['image_data'] = get_image_base64(
                                temp_question['image'].data)
                            
                            # Save resized dimensions from the form if they exist
                            resized_width = request.form.get(f'image_width_{question_copy.id}')
                            resized_height = request.form.get(f'image_height_{question_copy.id}')
                            if resized_width and resized_height:
                                temp_question['resized_width'] = resized_width
                                temp_question['resized_height'] = resized_height
                        
                        # Check for split questions
                        sub_q_id = request.form.get(f'partB_{q["qno"]}_sub')
                        if sub_q_id:
                            # Apply 'a)' label to main question
                            temp_question['sub_label'] = 'a)'
                            
                            # For split questions, maintain the full marks value for display
                            # This ensures both parts show the same mark value
                            temp_question['display_marks'] = marks_value
                            
                            # Add the main question to the list
                            part_b_questions.append(type('Question', (), temp_question))
                            
                            # Process sub-question
                            sub_question = Question.query.get(sub_q_id)
                            if sub_question and sub_question.marks > 2:
                                sub_marks = request.form.get(f'marks_{q["qno"]}_sub')
                                sub_marks_value = int(sub_marks) if sub_marks else q.get('marks', 8) // 2
                                
                                # Create a temporary object for the sub-question
                                temp_sub_question = {
                                    'id': sub_question.id,
                                    'unit': sub_question.unit,
                                    'question_text': sub_question.question_text,
                                    'marks': sub_marks_value,
                                    'display_marks': marks_value,  # Use full question marks to match part (a)
                                    'co': sub_question.co,
                                    'bl': sub_question.bl,
                                    'qno': q["qno"],  # Include the actual question number from template
                                    'pdf_qno': request.form.get(f'partB_qno_{q["qno"]}', q["qno"]), # Same number as parent for split questions
                                    'sub_label': 'b)',
                                    'image': sub_question.image
                                }
                                
                                # Add image data if available
                                if temp_sub_question['image']:
                                    temp_sub_question['image_data'] = get_image_base64(
                                        temp_sub_question['image'].data)
                                    
                                    # Save resized dimensions from the form if they exist
                                    resized_width = request.form.get(f'image_width_{sub_question.id}')
                                    resized_height = request.form.get(f'image_height_{sub_question.id}')
                                    if resized_width and resized_height:
                                        temp_sub_question['resized_width'] = resized_width
                                        temp_sub_question['resized_height'] = resized_height
                                
                                # Add the sub-question to the list
                                part_b_questions.append(type('Question', (), temp_sub_question))
                        else:
                            # No sub-question, add main question without label
                            part_b_questions.append(type('Question', (), temp_question))

            # Skip validation for Part A questions as requested
            # Just check if we have any Part A questions
            if len(part_a_questions) == 0:
                flash('Please select at least one Part A question', 'error')
                return redirect(request.url)

            if len(part_b_questions) < len(template.partB):
                print(
                    f"Part B: Expected {len(template.partB)}, got {len(part_b_questions)}"
                )
                flash('Please select all Part B questions', 'error')
                return redirect(request.url)

            # Check for preview mode - could be in args or form
            preview_mode = request.args.get('preview') or request.form.get(
                'preview')
            if preview_mode:
                # For preview, render the PDF template directly
                # Add image data to questions if available
                for question in part_a_questions + part_b_questions:
                    # Check if image attribute exists and has a value
                    if getattr(question, 'image', None):
                        question.image_data = f"data:image/png;base64,{get_image_base64(question.image.data)}"
                # Count part B main questions (excluding sub-questions) for the template calculation
                question_count = 0
                for question in part_b_questions:
                    if not hasattr(question, 'sub_label') or question.sub_label != 'b)':
                        question_count += 1
                
                # Get marks per question from actual questions
                if part_b_questions:
                    marks_per_question = part_b_questions[0].display_marks if hasattr(part_b_questions[0], 'display_marks') else part_b_questions[0].marks
                else:
                    marks_per_question = 8  # Default to 8 if no questions
                    
                # Prepare template data for PDF
                template_data = {
                    'header': template_header,
                    'partB': template.partB
                }
                
                return render_template('pdf_template.html',
                                       template=template_data,
                                       department=department,
                                       subject=subject,
                                       part_a_questions=part_a_questions,
                                       part_b_questions=part_b_questions,
                                       part_b_question_count=question_count,
                                       marks_per_question=marks_per_question,
                                       original_template=template)
            print(part_a_questions)

            # Save selected questions to the database for fuzzy logic tracking
            from models import SelectedQuestion

            # Track all the selected questions
            for question in part_a_questions + part_b_questions:
                # Check if this question has been selected before
                existing = SelectedQuestion.query.filter_by(
                    question_id=question.id,
                    subject_course_code=subject.course_code,
                    subject_department_id=subject.department_id,
                    exam_type=exam_type).first()

                if existing:
                    # Increment usage count
                    existing.usage_count += 1
                    existing.selection_date = datetime.datetime.utcnow(
                    )  # Update date
                else:
                    # Create new record
                    selected = SelectedQuestion(question_id=question.id,
                                                subject_course_code=subject.course_code,
                                                subject_department_id=subject.department_id,
                                                exam_type=exam_type)
                    db.session.add(selected)

            # Commit changes to database
            db.session.commit()

            # Get marks per question from actual questions
            if part_b_questions:
                marks_per_question = part_b_questions[0].display_marks if hasattr(part_b_questions[0], 'display_marks') else part_b_questions[0].marks
            else:
                marks_per_question = 8  # Default to 8 if no questions
                
            # Generate PDF
            pdf = generate_pdf(template_header, part_a_questions,
                               part_b_questions, department, subject,
                               selected_departments=selected_departments,
                               original_template=template)

            # Check if this is a preview request from AJAX call
            is_preview = request.form.get('preview') == 'true'

            # For preview, don't set as_attachment to allow in-page viewing
            if is_preview:
                return send_file(io.BytesIO(pdf),
                                mimetype='application/pdf',
                                as_attachment=False,
                                download_name='question_paper_preview.pdf')
            else:
                return send_file(io.BytesIO(pdf),
                                mimetype='application/pdf',
                                as_attachment=True,
                                download_name='Question_Paper.pdf')
        except Exception as e:
            flash(f'Error generating paper: {str(e)}', 'error')
            return redirect(
                url_for('generate_paper',
                        template_id=template_id,
                        course_code=course_code,
                        dept_id=dept_id,
                        exam_type=exam_type))

    return render_template('generate_paper.html',
                           template=template,
                           department=department,
                           subject=subject,
                           paper_type=paper_type,
                           related_departments=related_departments)


def get_questions_for_template(template):
    # Get questions for Part A (2 marks)
    partA_questions = []
    for q in template.partA:
        questions = Question.query.filter_by(unit=int(q['unit']),
                                             marks=2).limit(1).all()
        if questions:
            partA_questions.extend(questions)

    # Get questions for Part B (10 marks)
    partB_questions = []
    for q in template.partB:
        # Check if this is a split question (5+5 marks) or full question (10 marks)
        is_split = q.get('split', False)

        if is_split:
            # For split questions (a,b parts), get 5-mark questions
            questions = Question.query.filter((Question.unit == int(q['unit']))
                                              & (Question.marks == 5)).limit(
                                                  1).all()
        else:
            # For non-split questions, get 10-mark questions
            questions = Question.query.filter((Question.unit == int(q['unit']))
                                              & (Question.marks == 10)).limit(
                                                  1).all()

        if questions:
            partB_questions.extend(questions)

    return partA_questions + partB_questions


@app.route('/department/<string:dept_id>')
@login_required
def department_view(dept_id):
    department = Department.query.get_or_404(dept_id)
    regulations = Regulation.query.all()
    return render_template('department_view.html',
                           department=department,
                           regulations=regulations)


@app.route('/department/<string:dept_id>/analysis')
@login_required
def department_analysis(dept_id):
    """Display analysis of the question bank for a department"""
    department = Department.query.get_or_404(dept_id)
    subjects = Subject.query.filter_by(department_id=dept_id).all()

    # Get question data for all subjects in the department
    subject_data = []

    for subject in subjects:
        # Get all questions for this subject
        questions = Question.query.filter_by(
            subject_id=subject.id).all()

        # Count BL levels
        bl_counts = {}
        for question in questions:
            if question.bl not in bl_counts:
                bl_counts[question.bl] = 0
            bl_counts[question.bl] += 1

        # Count by units
        unit_counts = {}
        for question in questions:
            if question.unit not in unit_counts:
                unit_counts[question.unit] = 0
            unit_counts[question.unit] += 1

        # Get selected questions data
        selected_questions = SelectedQuestion.query.filter_by(
            subject_id=subject.id).all()
        exam_type_counts = {}
        for sq in selected_questions:
            if sq.exam_type not in exam_type_counts:
                exam_type_counts[sq.exam_type] = 0
            exam_type_counts[sq.exam_type] += 1

        # Add to subject data
        subject_data.append({
            'id': subject.id,
            'course_code': subject.course_code,
            'course_name': subject.course_name,
            'total_questions': len(questions),
            'bl_data': bl_counts,
            'unit_data': unit_counts,
            'exam_type_data': exam_type_counts
        })

    return render_template('department_analysis.html',
                           department=department,
                           subjects=subjects,
                           subject_data=subject_data)


@app.route('/department/<string:dept_id>/add_subject', methods=['POST', 'GET'])
@login_required
def add_subject(dept_id):
    department = Department.query.get_or_404(dept_id)

    if 'question_bank' not in request.files:
        flash('No file uploaded', 'error')
        return redirect(url_for('department_view', dept_id=dept_id))

    file = request.files['question_bank']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('department_view', dept_id=dept_id))

    if not file.filename.endswith('.xlsx'):
        flash('Please upload an Excel file', 'error')
        return redirect(url_for('department_view', dept_id=dept_id))

    try:
        subject = Subject(
            course_name=request.form['course_name'],
            course_code=request.form['course_code'],
            year=int(request.form['year']),
            semester=int(request.form['semester']),
            department_id=dept_id,
            regulation_id=request.form['regulation_id'],
        )
        db.session.add(subject)
        db.session.flush()

        # Load workbook
        workbook = load_workbook(file, data_only=True)

        # Process each worksheet (unit)
        for sheet_name in workbook.sheetnames:
            if not sheet_name.lower().startswith('unit'):
                continue

            worksheet = workbook[sheet_name]
            unit_number = int(''.join(filter(str.isdigit, sheet_name)))

            # Get header row to find column indices
            headers = {}
            for cell in worksheet[1]:
                if cell.value:
                    col_name = cell.value.strip().upper()
                    headers[col_name] = cell.column

            # Verify required columns exist
            required_columns = {'QUESTIONS', 'MARKS', 'CO', 'BL'}
            if not all(col in headers for col in required_columns):
                missing = [
                    col for col in required_columns if col not in headers
                ]
                flash(
                    f'Missing required columns in {sheet_name}: {", ".join(missing)}',
                    'error')
                continue

            # Process each row starting from row 2
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2),
                                          start=2):
                # Skip empty rows
                if not any(cell.value for cell in row):
                    continue

                # Get values from cells
                question_text = row[headers['QUESTIONS'] - 1].value
                marks = row[headers['MARKS'] - 1].value
                co = row[headers['CO'] - 1].value
                bl = row[headers['BL'] - 1].value

                # Skip if required fields are missing
                if not all([question_text, marks, co, bl]):
                    continue

                # Create question
                question = Question(unit=unit_number,
                                    question_text=str(question_text),
                                    marks=int(marks),
                                    co=str(co),
                                    bl=str(bl),
                                    subject_course_code=subject.course_code,
                                    subject_department_id=subject.department_id)
                db.session.add(question)
                db.session.flush()

                # Check for image in the row
                if 'IMAGE' in headers:
                    image_cell = row[headers['IMAGE'] - 1]
                    try:
                        image_data = extract_image_from_excel(
                            workbook, worksheet, image_cell)
                        if image_data:
                            image = Image(data=image_data,
                                          question_id=question.id)
                            db.session.add(image)
                    except Exception as e:
                        print(f"Error processing image: {e}")
                        continue

        db.session.commit()
        flash('Subject added successfully!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error adding subject: {str(e)}', 'error')

    return redirect(url_for('department_view', dept_id=dept_id))


@app.route('/subject/<string:course_code>/questions/<string:dept_id>')
@login_required
def view_question_bank(course_code, dept_id):
    subject = Subject.query.filter_by(course_code=course_code, department_id=dept_id).first_or_404()
    questions = Question.query.filter_by(subject_course_code=course_code, subject_department_id=dept_id).all()
    department = Department.query.get_or_404(dept_id)

    # Convert image data to base64 for display
    questions_with_images = []
    for question in questions:

        question_data = {
            'id': question.id,
            'unit': question.unit,
            'text': question.question_text,
            'marks': question.marks,
            'co': question.co,
            'bl': question.bl,
            'image_data': None
        }
        if question.image:
            image_base64 = get_image_base64(question.image.data)
            if image_base64:
                question_data['image_data'] = image_base64
        questions_with_images.append(question_data)

    return render_template('question_bank.html',
                           subject=subject,
                           questions=questions_with_images,
                           department=department)


@app.route('/subject/<string:subject_id>/delete/<string:dept_id>', methods=['POST'])
@login_required
def delete_subject(subject_id, dept_id):
    subject = Subject.query.filter_by(course_code=subject_id, department_id=dept_id).first_or_404()
    dept_id = subject.department_id
    course_code = subject.course_code

    try:
        # Delete selected questions related to the subject
        SelectedQuestion.query.filter_by(subject_course_code=subject.course_code, 
                                       subject_department_id=subject.department_id).delete()

        # Delete questions and their images
        questions = Question.query.filter_by(subject_course_code=subject.course_code, 
                                           subject_department_id=subject.department_id).all()
        for question in questions:
            if question.image:
                db.session.delete(question.image)
            db.session.delete(question)

        # Delete the subject
        db.session.delete(subject)
        db.session.commit()
        flash('Subject deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting subject: {str(e)}', 'error')

    return redirect(url_for('department_view', dept_id=dept_id))


@app.route('/department/<string:dept_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_department(dept_id):
    department = Department.query.get_or_404(dept_id)
    form = DepartmentForm(obj=department)

    if form.validate_on_submit():
        department.name = form.name.data
        department.dept_id = form.dept_id.data
        db.session.commit()
        flash('Department updated successfully!', 'success')
        return redirect(url_for('departments'))

    return render_template('department_form.html', form=form)


# Add delete department route
@app.route('/department/<string:dept_id>/delete', methods=['POST'])
@login_required
def delete_department(dept_id):
    department = Department.query.get_or_404(dept_id)
    try:
        db.session.delete(department)
        db.session.commit()
        flash('Department deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting department: {str(e)}', 'error')

    return redirect(url_for('departments'))


@app.route('/subject/<string:course_code>/edit/<string:dept_id>',
           methods=['GET', 'POST'])
@login_required
def edit_subject(course_code, dept_id):
    subject = Subject.query.filter_by(course_code=course_code, department_id=dept_id).first_or_404()
    department = Department.query.get_or_404(dept_id)
    regulations = Regulation.query.all()

    if request.method == 'POST':
        subject.course_name = request.form['course_name']
        subject.year = int(request.form['year'])
        subject.semester = int(request.form['semester'])
        subject.regulation_id = request.form['regulation_id']

        # Check if a new question bank file was uploaded
        if 'question_bank' in request.files and request.files[
                'question_bank'].filename:
            file = request.files['question_bank']

            if not file.filename.endswith('.xlsx'):
                flash('Please upload an Excel file', 'error')
                return redirect(request.url)

            try:
                # Delete old questions first
                old_questions = Question.query.filter_by(
                    subject_course_code=subject.course_code,
                    subject_department_id=subject.department_id).all()
                for question in old_questions:
                    if question.image:
                        db.session.delete(question.image)
                    db.session.delete(question)
                db.session.commit()

                # Load workbook
                workbook = load_workbook(file, data_only=True)

                # Process each worksheet (unit)
                for sheet_name in workbook.sheetnames:
                    if not sheet_name.lower().startswith('unit'):
                        continue

                    worksheet = workbook[sheet_name]
                    unit_number = int(''.join(filter(str.isdigit, sheet_name)))

                    # Find headers
                    headers = {}
                    for col_idx, cell in enumerate(worksheet[1], 1):
                        if cell.value:
                            header_text = str(cell.value).strip().upper()
                            if 'QUESTION' in header_text:
                                headers['QUESTION'] = col_idx
                            elif 'MARKS' in header_text:
                                headers['MARKS'] = col_idx
                            elif 'CO' in header_text:
                                headers['CO'] = col_idx
                            elif 'BL' in header_text:
                                headers['BL'] = col_idx
                            elif 'IMAGE' in header_text:
                                headers['IMAGE'] = col_idx

                    # Skip if required headers are missing
                    if not all(k in headers
                               for k in ['QUESTION', 'MARKS', 'CO', 'BL']):
                        continue

                    # Process rows
                    for row_idx, row in enumerate(
                            worksheet.iter_rows(min_row=2), 2):
                        # Skip empty rows
                        if all(cell.value is None for cell in row):
                            continue

                        question_cell = row[headers['QUESTION'] - 1]
                        marks_cell = row[headers['MARKS'] - 1]
                        co_cell = row[headers['CO'] - 1]
                        bl_cell = row[headers['BL'] - 1]

                        question_text = question_cell.value or ""
                        marks = marks_cell.value or 0
                        co = co_cell.value or ""
                        bl = bl_cell.value or ""

                        # Skip rows with empty required fields
                        if not question_text or not marks:
                            continue

                        # Create question
                        question = Question(unit=unit_number,
                                            question_text=str(question_text),
                                            marks=int(marks),
                                            co=str(co),
                                            bl=str(bl),
                                            subject_course_code=subject.course_code,
                                            subject_department_id=subject.department_id)
                        db.session.add(question)
                        db.session.flush()

                        # Check for image in the row
                        if 'IMAGE' in headers:
                            image_cell = row[headers['IMAGE'] - 1]
                            try:
                                image_data = extract_image_from_excel(
                                    workbook, worksheet, image_cell)
                                if image_data:
                                    image = Image(data=image_data,
                                                  question_id=question.id)
                                    db.session.add(image)
                            except Exception as e:
                                print(f"Error processing image: {e}")
                                continue

                db.session.commit()
                flash('Subject and question bank updated successfully!',
                      'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Error updating question bank: {str(e)}', 'error')
                return redirect(request.url)
        else:
            db.session.commit()
            flash('Subject updated successfully!', 'success')

        return redirect(url_for('department_view', dept_id=dept_id))

    return render_template('edit_subject.html',
                           subject=subject,
                           department=department,
                           regulations=regulations)


@app.route('/template/<int:template_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_template(template_id):
    template = Template.query.get_or_404(template_id)
    regulation = Regulation.query.get_or_404(template.regulation_id)

    if request.method == 'POST':
        try:
            # Update template name if provided
            if 'template_name' in request.form:
                template.name = request.form['template_name']
                
            # Update header information
            template.header = {
                'institution': request.form['institution'],
                'autonomous': request.form.get('autonomous', 'AUTONOMOUS'),
                'time_hours': int(request.form['time_hours']),
                'max_marks': int(request.form['max_marks'])
            }

            # Update Part A data
            try:
                partA_data = json.loads(request.form['partA'])
                if not partA_data:
                    flash('At least one Part A question is required', 'error')
                    return redirect(request.url)
                    
                # Make sure qno and sub_qno are preserved
                for i, question in enumerate(partA_data):
                    # Ensure we preserve the question number (always 1 for Part A)
                    question['qno'] = 1
                    # We set the sub-question letter based on index
                    question['sub_qno'] = f"{chr(97 + i)}"
                    
                template.partA = partA_data

                # Update Part B data
                partB_data = json.loads(request.form['partB'])
                if not partB_data:
                    flash('At least one Part B question is required', 'error')
                    return redirect(request.url)
                    
                # For Part B, ensure we preserve proper question numbering
                for i, question in enumerate(partB_data):
                    # Question number starts from 2 for Part B (each unit gets one pair of questions)
                    question['qno'] = i + 2
                    
                template.partB = partB_data
            except json.JSONDecodeError as e:
                flash(f'Error updating template: {str(e)}', 'error')
                return redirect(request.url)

            db.session.commit()
            flash('Template updated successfully!', 'success')
            return redirect(url_for('regulations'))

        except Exception as e:
            flash(f'Error updating template: {str(e)}', 'error')
            return redirect(request.url)

    return render_template('edit_template.html',
                           template=template,
                           regulation=regulation,
                           exam_type=template.exam_type,
                           is_edit=True)


@app.route('/get_filtered_chart_data')
@login_required
def get_filtered_chart_data():
    """Get filtered chart data based on selected filters"""
    # Get filter parameters
    department_id = request.args.get('department', 'all')
    regulation_id = request.args.get('regulation', 'all')
    exam_type = request.args.get('exam_type', 'all')
    paper_type = request.args.get('paper_type', 'all')
    
    # Build query for selected questions
    query = SelectedQuestion.query
    
    # Apply filters
    if department_id != 'all':
        query = query.filter_by(subject_department_id=department_id)
    
    if exam_type != 'all':
        query = query.filter_by(exam_type=exam_type)
    
    # Get subjects that match regulation filter
    if regulation_id != 'all':
        # Get all subjects with this regulation
        regulation_subjects = Subject.query.filter_by(regulation_id=regulation_id).all()
        subject_codes = [(s.course_code, s.department_id) for s in regulation_subjects]
        
        # Filter selected questions to only include those with matching subject codes
        filtered_questions = []
        for sq in query.all():
            if (sq.subject_course_code, sq.subject_department_id) in subject_codes:
                filtered_questions.append(sq)
    else:
        filtered_questions = query.all()
    
    # Initialize data structures for chart
    co_distribution = {}
    bl_distribution = {}
    
    for sq in filtered_questions:
        # Get the actual question
        question = Question.query.get(sq.question_id)
        if question:
            # Strip whitespace from CO and BL values to prevent duplicates
            co_value = question.co.strip() if isinstance(question.co, str) else question.co
            bl_value = question.bl.strip() if isinstance(question.bl, str) else question.bl
            
            # CO distribution
            if co_value in co_distribution:
                co_distribution[co_value] += 1
            else:
                co_distribution[co_value] = 1
                
            # BL distribution
            if bl_value in bl_distribution:
                bl_distribution[bl_value] += 1
            else:
                bl_distribution[bl_value] = 1
    
    # Convert to format needed for Chart.js
    co_labels = list(co_distribution.keys())
    co_values = list(co_distribution.values())
    
    bl_labels = list(bl_distribution.keys())
    bl_values = list(bl_distribution.values())
    
    return jsonify({
        'co_labels': co_labels,
        'co_values': co_values,
        'bl_labels': bl_labels,
        'bl_values': bl_values,
        'co_total': sum(co_values) if co_values else 0,
        'bl_total': sum(bl_values) if bl_values else 0
    })


@app.route('/subject/<string:course_code>/analysis/<string:dept_id>')
@login_required
def subject_analysis(course_code, dept_id):
    subject = Subject.query.filter_by(course_code=course_code, department_id=dept_id).first_or_404()
    questions = Question.query.filter_by(subject_course_code=course_code, subject_department_id=dept_id).all()

    # Get BL level data
    bl_data = {}
    for question in questions:
        # Strip whitespace from BL value to prevent duplicates
        bl_value = question.bl.strip() if isinstance(question.bl, str) else question.bl
        if bl_value not in bl_data:
            bl_data[bl_value] = 0
        bl_data[bl_value] += 1

    # Get unit-wise BL data
    unit_bl_data = {}
    for question in questions:
        # Strip whitespace from BL value to prevent duplicates
        bl_value = question.bl.strip() if isinstance(question.bl, str) else question.bl
        
        if question.unit not in unit_bl_data:
            unit_bl_data[question.unit] = {}
        if bl_value not in unit_bl_data[question.unit]:
            unit_bl_data[question.unit][bl_value] = 0
        unit_bl_data[question.unit][bl_value] += 1

    # Get paper analysis data
    paper_data = {}
    selected_questions = SelectedQuestion.query.filter_by(
        subject_course_code=course_code, subject_department_id=dept_id).all()
    for sq in selected_questions:
        exam_type = sq.exam_type
        if exam_type not in paper_data:
            paper_data[exam_type] = {}
            
        # Strip whitespace from BL value to prevent duplicates
        bl_value = sq.question.bl.strip() if isinstance(sq.question.bl, str) else sq.question.bl
        
        if bl_value not in paper_data[exam_type]:
            paper_data[exam_type][bl_value] = 0
        paper_data[exam_type][bl_value] += 1

    # Convert questions to dict for JavaScript
    questions_list = [{'bl': q.bl, 'marks': q.marks} for q in questions]

    return render_template('subject_analysis.html',
                           subject=subject,
                           bl_data=bl_data,
                           unit_bl_data=unit_bl_data,
                           paper_data=paper_data,
                           questions_by_marks=questions_list)
